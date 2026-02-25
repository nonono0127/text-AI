#!/usr/bin/env python3
"""
所見生成ツール - ExcelファイルをドロップしてE列に所見を自動生成するGUIアプリ

使い方:
    python shoken_app.py

必要なライブラリ:
    pip install anthropic openpyxl xlrd>=2.0.1 tkinterdnd2
"""

import os
import sys
import threading
import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext, ttk

import re

import anthropic
import openpyxl
from openpyxl.styles import Alignment

# ドラッグ＆ドロップサポート（任意）
try:
    from tkinterdnd2 import DND_FILES, TkinterDnD
    HAS_DND = True
except ImportError:
    HAS_DND = False

# .xls 読み込みサポート（任意）
try:
    import xlrd
    HAS_XLRD = True
except ImportError:
    HAS_XLRD = False

# 所見テンプレートを読み込む
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from templates import TEMPLATES

SHOKEN = TEMPLATES["shoken"]


def calc_length_range(target_length_str: str) -> tuple[int, int]:
    """'170字' → (155, 185) のように目標文字数から±15の範囲を返す"""
    m = re.search(r'\d+', str(target_length_str))
    target = int(m.group()) if m else 170
    return target - 15, target + 15


class ShokenApp:
    def __init__(self, root):
        self.root = root
        self.root.title("所見生成ツール")
        self.root.geometry("700x660")
        self.root.resizable(True, True)
        self.file_path = None
        self.is_running = False
        self._build_ui()

    # ------------------------------------------------------------------ UI --

    def _build_ui(self):
        # ヘッダー
        hdr = tk.Frame(self.root, bg="#2c3e50", padx=10, pady=10)
        hdr.pack(fill=tk.X)
        tk.Label(hdr, text="所見生成ツール",
                 font=("Arial", 18, "bold"), bg="#2c3e50", fg="white").pack()
        tk.Label(hdr,
                 text="ExcelファイルをドロップするとE列に所見を自動生成します",
                 font=("Arial", 9), bg="#2c3e50", fg="#bdc3c7").pack()

        main = tk.Frame(self.root, padx=12, pady=8)
        main.pack(fill=tk.BOTH, expand=True)

        # --- Excelファイル選択 ---
        file_lf = tk.LabelFrame(main, text="Excelファイル", padx=8, pady=8)
        file_lf.pack(fill=tk.X, pady=(0, 8))

        if HAS_DND:
            self.drop_lbl = tk.Label(
                file_lf,
                text="ここにExcelファイルをドラッグ＆ドロップ",
                font=("Arial", 12), relief=tk.SUNKEN,
                padx=10, pady=24, bg="#ecf0f1", cursor="hand2"
            )
            self.drop_lbl.pack(fill=tk.X, pady=(0, 6))
            self.drop_lbl.drop_target_register(DND_FILES)
            self.drop_lbl.dnd_bind("<<Drop>>", self._on_drop)
        else:
            tk.Label(
                file_lf,
                text="【注意】tkinterdnd2 が未インストールのためドラッグ＆ドロップ不可。\n"
                     "下の「参照...」ボタンからファイルを選択してください。",
                font=("Arial", 9), fg="#e74c3c", justify=tk.LEFT
            ).pack(anchor=tk.W, pady=(0, 6))

        path_row = tk.Frame(file_lf)
        path_row.pack(fill=tk.X)
        self.file_var = tk.StringVar(value="ファイルが選択されていません")
        tk.Entry(path_row, textvariable=self.file_var, state="readonly",
                 font=("Arial", 9)).pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 4))
        tk.Button(path_row, text="参照...", command=self._browse,
                  bg="#3498db", fg="white", padx=8).pack(side=tk.RIGHT)

        # --- 入力列の説明 ---
        fmt_lf = tk.LabelFrame(main, text="入力Excelの列構成（1行目はヘッダー推奨）", padx=8, pady=6)
        fmt_lf.pack(fill=tk.X, pady=(0, 8))
        cols = [
            ("A列", "学年", "例: 3年生"),
            ("B列", "係・委員会・行事での様子", "例: 給食係として…"),
            ("C列", "学習面での様子", "例: 算数の授業では…"),
            ("D列", "文字数目安", "例: 170字"),
            ("E列", "★生成された所見が書き込まれます", ""),
        ]
        for col, label, hint in cols:
            row_f = tk.Frame(fmt_lf)
            row_f.pack(fill=tk.X)
            tk.Label(row_f, text=col, width=5, font=("Arial", 9, "bold"),
                     fg="#2980b9", anchor=tk.W).pack(side=tk.LEFT)
            tk.Label(row_f, text=label, font=("Arial", 9), anchor=tk.W).pack(side=tk.LEFT)
            if hint:
                tk.Label(row_f, text=f"  {hint}", font=("Arial", 8),
                         fg="#7f8c8d", anchor=tk.W).pack(side=tk.LEFT)

        # --- 設定 ---
        cfg_lf = tk.LabelFrame(main, text="設定", padx=8, pady=8)
        cfg_lf.pack(fill=tk.X, pady=(0, 8))

        api_row = tk.Frame(cfg_lf)
        api_row.pack(fill=tk.X, pady=(0, 4))
        tk.Label(api_row, text="APIキー:", width=14, anchor=tk.W).pack(side=tk.LEFT)
        self.api_var = tk.StringVar(value=os.environ.get("ANTHROPIC_API_KEY", ""))
        tk.Entry(api_row, textvariable=self.api_var, show="*",
                 font=("Arial", 9)).pack(side=tk.LEFT, fill=tk.X, expand=True)

        row_row = tk.Frame(cfg_lf)
        row_row.pack(fill=tk.X, pady=(0, 4))
        tk.Label(row_row, text="開始行:", width=14, anchor=tk.W).pack(side=tk.LEFT)
        self.start_row_var = tk.StringVar(value="2")
        tk.Entry(row_row, textvariable=self.start_row_var, width=5).pack(side=tk.LEFT)
        tk.Label(row_row, text=" (1行目がヘッダーの場合は2)",
                 fg="#7f8c8d", font=("Arial", 8)).pack(side=tk.LEFT)

        self.overwrite_var = tk.BooleanVar(value=False)
        tk.Checkbutton(cfg_lf, text="E列に既存データがある行も上書きする",
                       variable=self.overwrite_var).pack(anchor=tk.W)

        # --- 生成ボタン ---
        self.gen_btn = tk.Button(
            main, text="所見を生成する", command=self._start,
            font=("Arial", 13, "bold"), bg="#27ae60", fg="white",
            padx=16, pady=6, state=tk.DISABLED
        )
        self.gen_btn.pack(pady=(0, 8))

        # --- 進捗 ---
        prog_lf = tk.LabelFrame(main, text="進捗", padx=8, pady=6)
        prog_lf.pack(fill=tk.X, pady=(0, 8))
        self.prog_var = tk.DoubleVar()
        ttk.Progressbar(prog_lf, variable=self.prog_var,
                        maximum=100).pack(fill=tk.X, pady=(0, 4))
        self.status_var = tk.StringVar(value="待機中")
        tk.Label(prog_lf, textvariable=self.status_var,
                 font=("Arial", 9), anchor=tk.W).pack(fill=tk.X)

        # --- ログ ---
        log_lf = tk.LabelFrame(main, text="ログ", padx=4, pady=4)
        log_lf.pack(fill=tk.BOTH, expand=True)
        self.log = scrolledtext.ScrolledText(
            log_lf, height=7, font=("Courier", 9), state=tk.DISABLED)
        self.log.pack(fill=tk.BOTH, expand=True)

    # --------------------------------------------------------------- helpers --

    def _log(self, msg):
        self.log.config(state=tk.NORMAL)
        self.log.insert(tk.END, msg + "\n")
        self.log.see(tk.END)
        self.log.config(state=tk.DISABLED)

    def _on_drop(self, event):
        path = event.data.strip()
        # Windows では {} で囲まれることがある
        if path.startswith("{") and path.endswith("}"):
            path = path[1:-1]
        self._load_file(path)

    def _browse(self):
        path = filedialog.askopenfilename(
            title="Excelファイルを選択",
            filetypes=[("Excelファイル", "*.xlsx *.xls"), ("すべてのファイル", "*.*")]
        )
        if path:
            self._load_file(path)

    def _load_file(self, path):
        if not os.path.exists(path):
            messagebox.showerror("エラー", f"ファイルが見つかりません:\n{path}")
            return
        ext = os.path.splitext(path)[1].lower()
        if ext == ".xls" and not HAS_XLRD:
            messagebox.showerror(
                "エラー",
                ".xlsファイルを読み込むには xlrd が必要です。\n"
                "pip install xlrd>=2.0.1 を実行してください。"
            )
            return
        if ext not in (".xlsx", ".xls"):
            messagebox.showerror("エラー", "XLSX または XLS 形式のExcelファイルを選択してください")
            return
        self.file_path = path
        self.file_var.set(path)
        self.gen_btn.config(state=tk.NORMAL)
        self._log(f"ファイル選択: {os.path.basename(path)}")
        if HAS_DND:
            self.drop_lbl.config(text=f"✓ {os.path.basename(path)}", bg="#d5f5e3")

    # ----------------------------------------------------------- generation --

    def _start(self):
        if self.is_running:
            return
        api_key = self.api_var.get().strip()
        if not api_key:
            messagebox.showerror("エラー", "APIキーを入力してください")
            return
        if not self.file_path:
            messagebox.showerror("エラー", "Excelファイルを選択してください")
            return
        try:
            start_row = int(self.start_row_var.get())
        except ValueError:
            messagebox.showerror("エラー", "開始行は数値で入力してください")
            return

        self.is_running = True
        self.gen_btn.config(state=tk.DISABLED, text="生成中...")
        self.prog_var.set(0)
        threading.Thread(
            target=self._worker,
            args=(api_key, start_row, self.overwrite_var.get()),
            daemon=True
        ).start()

    def _worker(self, api_key, start_row, overwrite):
        try:
            wb, ws, rows_data = self._read_excel(start_row)
            total = len(rows_data)
            if total == 0:
                self.root.after(0, lambda: messagebox.showinfo(
                    "情報",
                    "処理対象の行が見つかりませんでした。\n"
                    "A〜D列にデータが入力されているか確認してください。"
                ))
                return

            self.root.after(0, lambda: self._log(f"処理対象: {total} 行"))
            client = anthropic.Anthropic(api_key=api_key)
            ok = skip = err = 0

            for i, row in enumerate(rows_data):
                ridx = row["row_idx"]

                if row["existing_e"] and not overwrite:
                    self.root.after(
                        0, lambda r=ridx: self._log(f"  行{r}: スキップ（E列に既存データあり）"))
                    skip += 1
                else:
                    self.root.after(
                        0, lambda r=ridx, n=i + 1, t=total:
                        self.status_var.set(f"行{r} を処理中... ({n}/{t})"))
                    self.root.after(0, lambda r=ridx: self._log(f"  行{r}: 生成中..."))
                    try:
                        text = self._call_api(client, row)
                        ws.cell(ridx, 5, text).alignment = Alignment(wrap_text=True)
                        clen = len(text)
                        self.root.after(
                            0, lambda r=ridx, c=clen: self._log(f"  行{r}: 完了 ({c}字)"))
                        ok += 1
                    except Exception as e:
                        emsg = str(e)
                        self.root.after(
                            0, lambda r=ridx, e=emsg: self._log(f"  行{r}: エラー - {e}"))
                        err += 1

                pct = (i + 1) / total * 100
                self.root.after(0, lambda p=pct: self.prog_var.set(p))

            # 保存
            base = os.path.splitext(self.file_path)[0]
            out_path = base + "_生成済み.xlsx"
            wb.save(out_path)

            summary = f"完了: {ok}件生成 / {skip}件スキップ / {err}件エラー"
            self.root.after(0, lambda: self.status_var.set(summary))
            self.root.after(0, lambda: self._log(f"\n{summary}"))
            self.root.after(0, lambda: self._log(f"保存先: {out_path}"))
            self.root.after(0, lambda: messagebox.showinfo(
                "完了", f"{summary}\n\n保存先:\n{out_path}"))

        except anthropic.AuthenticationError:
            self.root.after(0, lambda: messagebox.showerror(
                "APIキーエラー",
                "APIキーが無効です。\n"
                "https://console.anthropic.com/settings/keys で確認してください。"
            ))
            self.root.after(0, lambda: self._log("エラー: APIキーが無効です"))
        except Exception as e:
            emsg = str(e)
            self.root.after(
                0, lambda: messagebox.showerror("エラー", f"処理中にエラーが発生しました:\n{emsg}"))
            self.root.after(0, lambda: self._log(f"エラー: {emsg}"))
        finally:
            self.is_running = False
            self.root.after(
                0, lambda: self.gen_btn.config(state=tk.NORMAL, text="所見を生成する"))

    def _read_excel(self, start_row):
        """Excelを読み込んでopenpyxlのWorkbook/Worksheet/行データリストを返す"""
        ext = os.path.splitext(self.file_path)[1].lower()
        if ext == ".xlsx":
            wb = openpyxl.load_workbook(self.file_path)
            ws = wb.active
        else:
            # .xls → openpyxl の Workbook に変換して処理
            xls = xlrd.open_workbook(self.file_path)
            xsh = xls.sheet_by_index(0)
            wb = openpyxl.Workbook()
            ws = wb.active
            for r in range(xsh.nrows):
                for c in range(xsh.ncols):
                    ws.cell(r + 1, c + 1, xsh.cell(r, c).value)

        rows_data = []
        for r in range(start_row, ws.max_row + 1):
            grade = ws.cell(r, 1).value
            activities = ws.cell(r, 2).value
            subject = ws.cell(r, 3).value
            target_len = ws.cell(r, 4).value
            existing_e = ws.cell(r, 5).value
            if grade or activities or subject:
                rows_data.append({
                    "row_idx": r,
                    "grade": str(grade or "3年生"),
                    "activities": str(activities or ""),
                    "subject_learning": str(subject or ""),
                    "target_length": str(target_len or "170字"),
                    "existing_e": existing_e,
                })
        return wb, ws, rows_data

    def _call_api(self, client, row):
        """Claude APIを呼び出して所見文を生成して返す"""
        min_len, max_len = calc_length_range(row["target_length"])
        system_prompt = SHOKEN["system_prompt"].format(
            target_length=row["target_length"],
            min_length=min_len,
            max_length=max_len,
        )
        user_prompt = SHOKEN["user_prompt_template"].format(
            grade=row["grade"],
            activities=row["activities"],
            subject_learning=row["subject_learning"],
            target_length=row["target_length"],
            min_length=min_len,
            max_length=max_len,
        )
        generated = []
        with client.messages.stream(
            model="claude-opus-4-6",
            max_tokens=600,
            system=system_prompt,
            messages=[{"role": "user", "content": user_prompt}],
        ) as stream:
            for text in stream.text_stream:
                generated.append(text)
        return "".join(generated).strip()


def main():
    if HAS_DND:
        root = TkinterDnD.Tk()
    else:
        root = tk.Tk()
    ShokenApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
