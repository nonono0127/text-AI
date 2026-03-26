#!/usr/bin/env python3
"""
所見生成ツール - Streamlit版

使い方:
    streamlit run streamlit_app.py
"""

import io
import os
import re
import secrets
import smtplib
import ssl
import sys
import time

import anthropic
import openpyxl
from openpyxl.styles import Alignment
import streamlit as st

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from templates import TEMPLATES

SHOKEN = TEMPLATES["shoken"]


def send_otp_email(to_email: str, otp: str) -> None:
    smtp_email = st.secrets["SMTP_EMAIL"]
    smtp_password = st.secrets["SMTP_PASSWORD"]
    subject = "所見生成ツール ワンタイムパスワード"
    body = f"ワンタイムパスワード: {otp}\n\n有効期限は10分です。"
    message = f"Subject: {subject}\nTo: {to_email}\n\n{body}"
    context = ssl.create_default_context()
    with smtplib.SMTP_SSL("smtp.gmail.com", 465, context=context) as server:
        server.login(smtp_email, smtp_password)
        server.sendmail(smtp_email, to_email, message.encode("utf-8"))


def auth_screen() -> bool:
    if st.session_state.get("authenticated"):
        return True

    st.title("所見生成ツール")
    allowed_emails = list(st.secrets.get("ALLOWED_EMAILS", []))

    if st.session_state.get("otp_sent"):
        st.info(f"{st.session_state.otp_email} にワンタイムパスワードを送信しました。")
        otp_input = st.text_input("ワンタイムパスワード（6桁）")
        col1, col2 = st.columns(2)
        with col1:
            if st.button("確認", use_container_width=True):
                if time.time() > st.session_state.otp_expiry:
                    st.error("有効期限が切れました。もう一度メールアドレスを入力してください。")
                    st.session_state.otp_sent = False
                    st.rerun()
                elif otp_input == st.session_state.otp_code:
                    st.session_state.authenticated = True
                    st.rerun()
                else:
                    st.error("パスワードが正しくありません。")
        with col2:
            if st.button("メールアドレスを変更", use_container_width=True):
                st.session_state.otp_sent = False
                st.rerun()
    else:
        email = st.text_input("メールアドレス")
        if st.button("送信", use_container_width=True):
            if email in allowed_emails:
                otp = str(secrets.randbelow(900000) + 100000)
                try:
                    send_otp_email(email, otp)
                    st.session_state.otp_code = otp
                    st.session_state.otp_email = email
                    st.session_state.otp_expiry = time.time() + 600
                    st.session_state.otp_sent = True
                    st.rerun()
                except Exception as e:
                    st.error(f"メール送信に失敗しました: {e}")
            else:
                st.error("このメールアドレスは登録されていません。")
    return False


def read_excel_from_bytes(file_bytes: bytes, filename: str, start_row: int):
    """Excelバイト列を読み込んでWorkbook/Worksheet/行データリストを返す"""
    ext = os.path.splitext(filename)[1].lower()

    if ext == ".xlsx":
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes))
        ws = wb.active
    elif ext == ".xls":
        try:
            import xlrd
        except ImportError:
            st.error(".xlsファイルには xlrd が必要です: pip install xlrd>=2.0.1")
            return None, None, []
        xls = xlrd.open_workbook(file_contents=file_bytes)
        xsh = xls.sheet_by_index(0)
        wb = openpyxl.Workbook()
        ws = wb.active
        for r in range(xsh.nrows):
            for c in range(xsh.ncols):
                ws.cell(r + 1, c + 1, xsh.cell(r, c).value)
    else:
        st.error("xlsx または xls 形式のファイルを選択してください。")
        return None, None, []

    rows_data = []
    for r in range(start_row, ws.max_row + 1):
        grade = ws.cell(r, 1).value
        activities = ws.cell(r, 2).value
        subject = ws.cell(r, 3).value
        target_len = ws.cell(r, 4).value
        existing_e = ws.cell(r, 5).value
        if grade or activities or subject:
            rows_data.append({
                "row_idx": r,
                "grade": str(grade or "3年生"),
                "activities": str(activities or ""),
                "subject_learning": str(subject or ""),
                "target_length": str(target_len or "170字"),
                "existing_e": existing_e,
            })
    return wb, ws, rows_data


def calc_length_range(target_length_str: str) -> tuple[int, int]:
    """'170字' → (155, 185) のように目標文字数から±15の範囲を返す"""
    m = re.search(r'\d+', str(target_length_str))
    target = int(m.group()) if m else 170
    return target - 15, target + 15


def call_api(client: anthropic.Anthropic, row: dict) -> str:
    """Claude APIを呼び出して所見文を生成して返す"""
    min_len, max_len = calc_length_range(row["target_length"])
    system_prompt = SHOKEN["system_prompt"].format(
        target_length=row["target_length"],
        min_length=min_len,
        max_length=max_len,
    )
    user_prompt = SHOKEN["user_prompt_template"].format(
        grade=row["grade"],
        activities=row["activities"],
        subject_learning=row["subject_learning"],
        target_length=row["target_length"],
        min_length=min_len,
        max_length=max_len,
    )
    generated = []
    with client.messages.stream(
        model="claude-opus-4-6",
        max_tokens=600,
        system=system_prompt,
        messages=[{"role": "user", "content": user_prompt}],
    ) as stream:
        for text in stream.text_stream:
            generated.append(text)
    return "".join(generated).strip()


def main():
    st.set_page_config(page_title="所見生成ツール", page_icon="📝", layout="centered")

    if not auth_screen():
        return

    api_key = st.secrets.get("ANTHROPIC_API_KEY", os.environ.get("ANTHROPIC_API_KEY", ""))
    if not api_key:
        st.error("APIキーが設定されていません。管理者に連絡してください。")
        return

    st.title("📝 所見生成ツール")
    st.caption("ExcelファイルをアップロードするとE列に通知表の所見を自動生成します")

    st.divider()

    # --- Excelフォーマット説明 ---
    with st.expander("📋 入力Excelの列構成", expanded=False):
        st.markdown("""
| 列 | 内容 | 例 |
|---|---|---|
| **A列** | 学年 | 3年生 |
| **B列** | 係・委員会・行事での様子 | 給食係として毎日丁寧に… |
| **C列** | 学習面での様子 | 算数の授業では… |
| **D列** | 文字数目安 | 170字 |
| **E列** | ★生成された所見が書き込まれます | （自動入力） |

1行目はヘッダー行にして、2行目から児童データを入力してください。
        """)

    # --- ファイルアップロード ---
    uploaded_file = st.file_uploader(
        "Excelファイルをアップロード",
        type=["xlsx", "xls"],
        help="ドラッグ＆ドロップまたはクリックして選択",
    )

    # --- 設定 ---
    col1, col2 = st.columns(2)
    with col1:
        start_row = st.number_input("開始行", min_value=1, value=2, step=1,
                                    help="1行目がヘッダーの場合は2")
    with col2:
        overwrite = st.checkbox("E列の既存データも上書きする", value=False)

    st.divider()

    # --- 生成ボタン ---
    generate_btn = st.button(
        "所見を生成する",
        type="primary",
        disabled=(uploaded_file is None or not api_key),
        use_container_width=True,
    )

    if uploaded_file is None:
        st.info("Excelファイルをアップロードしてください。")
        return

    if not api_key:
        st.error("APIキーが設定されていません。")
        return

    if not generate_btn:
        return

    # --- 生成処理 ---
    file_bytes = uploaded_file.read()
    wb, ws, rows_data = read_excel_from_bytes(file_bytes, uploaded_file.name, start_row)

    if wb is None:
        return

    total = len(rows_data)
    if total == 0:
        st.warning("処理対象の行が見つかりませんでした。A〜D列にデータが入力されているか確認してください。")
        return

    st.info(f"処理対象: {total} 行")

    try:
        client = anthropic.Anthropic(api_key=api_key)
    except Exception as e:
        st.error(f"APIクライアントの初期化に失敗しました: {e}")
        return

    progress_bar = st.progress(0)
    status_text = st.empty()
    log_area = st.empty()
    log_lines = []

    ok = skip = err = 0

    for i, row in enumerate(rows_data):
        ridx = row["row_idx"]

        if row["existing_e"] and not overwrite:
            log_lines.append(f"行{ridx}: スキップ（E列に既存データあり）")
            skip += 1
        else:
            status_text.text(f"行{ridx} を処理中... ({i + 1}/{total})")
            log_lines.append(f"行{ridx}: 生成中...")
            log_area.text("\n".join(log_lines[-20:]))

            try:
                text = call_api(client, row)
                ws.cell(ridx, 5, text).alignment = Alignment(wrap_text=True)
                clen = len(text)
                log_lines.append(f"行{ridx}: 完了 ({clen}字)")
                ok += 1
            except anthropic.AuthenticationError:
                st.error("APIキーが無効です。正しいキーを入力してください。")
                return
            except Exception as e:
                log_lines.append(f"行{ridx}: エラー - {e}")
                err += 1

        progress_bar.progress((i + 1) / total)
        log_area.text("\n".join(log_lines[-20:]))

    status_text.empty()
    log_area.text("\n".join(log_lines))

    # --- 結果サマリー ---
    summary = f"完了: {ok}件生成 / {skip}件スキップ / {err}件エラー"
    if err == 0:
        st.success(summary)
    else:
        st.warning(summary)

    # --- ダウンロード ---
    out_buf = io.BytesIO()
    wb.save(out_buf)
    out_buf.seek(0)

    base_name = os.path.splitext(uploaded_file.name)[0]
    out_filename = base_name + "_生成済み.xlsx"

    st.download_button(
        label="📥 生成済みExcelをダウンロード",
        data=out_buf,
        file_name=out_filename,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
        type="primary",
    )


if __name__ == "__main__":
    main()
